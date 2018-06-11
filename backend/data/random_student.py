from openpyxl import Workbook
import random
import string

# 姓名	学校	年级	学号	专业	电话号码	电子邮箱	账号状态

major = ['物理','化学','英语','计算机','经管','人文','环境','土木','机械','设计']

mail = ['163','162']

status = [True,False]

def random_phone():
    head = ['139','188','185','136','155','135','158','151','152','153']
    s = '0123456789'
    return random.choice(head)+"".join(random.choice(s) for i in range(8))

def main():
    wb = Workbook()
    ws = wb.create_sheet("学生信息")
    
    for i in range(1,200):
        name = ''.join(random.sample(string.ascii_letters + string.digits, 8))
        school_id = random.randint(1,6)
        grade = random.randint(2017,2020)
        code = '{:0>7}'.format(str(i))
        major_id = random.randint(0,len(major)-1)
        phone = random_phone()
        mail_name = phone+"@"+random.choice(mail)+".com"
        statuss = random.choice(status)
        print(name,school_id,grade,code,major[major_id],phone,mail_name,statuss)
        ws.cell(row=i,column=1).value = name
        ws.cell(row=i,column=2).value = school_id
        ws.cell(row=i,column=3).value = grade
        ws.cell(row=i,column=4).value = code
        ws.cell(row=i,column=5).value = major[major_id]
        ws.cell(row=i,column=6).value = phone
        ws.cell(row=i,column=7).value = mail_name
        ws.cell(row=i,column=8).value = statuss
        
    wb.save("random.xlsx")


if __name__ == '__main__':
    main()